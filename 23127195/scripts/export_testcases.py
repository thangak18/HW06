#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
export_testcases.py — Xuat test case ra dinh dang bang (Excel / CSV)

Doc IR trong `testcases/*.json` va doi chieu voi ket qua Newman moi nhat trong
`newman/*.json` de sinh:

  testcases/TESTCASES_23127195.xlsx   (neu co openpyxl)
  testcases/api{1,2,3}_testcases.csv  (luon luon)
  testcases/TEST_SUMMARY.md           (bang tong hop cho bao cao)

Cot xuat ra bam theo yeu cau cua de bai: moi test case deu co ky thuat ap dung,
tham so / phan vung, ky vong theo dac ta, NGUON (AI / HUMAN), NHAN AUDIT
(VALID / INVALID / INCOMPLETE) kem ly do, ket qua chay that va ma loi lien quan.

Usage:
    python scripts/export_testcases.py
"""

import csv
import glob
import io
import json
import os
import sys
from collections import Counter, OrderedDict

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
TC_DIR = os.path.join(ROOT, "testcases")
NEWMAN_DIR = os.path.join(ROOT, "newman")

COLUMNS = [
    ("id", "Test Case ID"),
    ("api", "API"),
    ("fr", "FR"),
    ("group", "Nhom"),
    ("title", "Tieu de test case"),
    ("technique", "Ky thuat"),
    ("param", "Tham so"),
    ("partition", "Phan vung / Gia tri bien"),
    ("sec", "SEC lien quan"),
    ("method", "Method"),
    ("endpoint", "Endpoint"),
    ("request_body", "Du lieu gui len"),
    ("expected_by_spec", "Ket qua ky vong (theo dac ta)"),
    ("source", "Nguon (AI/HUMAN)"),
    ("audit_label", "Nhan audit"),
    ("audit_reason", "Ly do audit / hieu chinh"),
    ("actual", "Ket qua chay that (Newman)"),
    ("verdict", "PASS/FAIL"),
    ("bug_id", "Ma loi"),
]


def load_newman_results():
    """Doc bao cao Newman moi nhat cua tung API -> {ten_request: (verdict, chi_tiet)}."""
    results = {}
    for key in ("api1", "api2", "api3"):
        files = sorted(glob.glob(os.path.join(NEWMAN_DIR, "%s_*.json" % key)))
        if not files:
            continue
        with io.open(files[-1], encoding="utf-8") as f:
            data = json.load(f)
        fails = {}
        for fail in data["run"].get("failures", []):
            name = fail.get("source", {}).get("name", "")
            fails.setdefault(name, []).append(fail["error"]["message"])
        for ex in data["run"].get("executions", []):
            name = ex["item"]["name"]
            tc_id = name.split()[0]
            n_assert = len(ex.get("assertions", []))
            if name in fails:
                results[tc_id] = ("FAIL", " | ".join(m[:160] for m in fails[name][:3]))
            elif n_assert:
                results[tc_id] = ("PASS", "%d assertion deu dat" % n_assert)
            else:
                results.setdefault(tc_id, ("N/A", "khong ghi nhan assertion"))
        results["__file__" + key] = (os.path.basename(files[-1]), "")
    return results


def flatten(spec, case, results):
    req = case.get("request", {})
    body = req.get("body", "")
    if isinstance(body, (dict, list)):
        body = json.dumps(body, ensure_ascii=False)
    body = str(body)
    if len(body) > 300:
        body = body[:297] + "..."
    verdict, actual = results.get(case["id"], ("CHUA CHAY", ""))
    return {
        "id": case["id"],
        "api": spec["api_id"],
        "fr": spec["fr"],
        "group": case.get("group", ""),
        "title": case["title"],
        "technique": case.get("technique", ""),
        "param": case.get("param", ""),
        "partition": case.get("partition", ""),
        "sec": ", ".join(case.get("sec", [])),
        "method": req.get("method", ""),
        "endpoint": req.get("path", ""),
        "request_body": body,
        "expected_by_spec": case.get("expected_by_spec", ""),
        "source": case.get("source", ""),
        "audit_label": case.get("audit", {}).get("label", ""),
        "audit_reason": case.get("audit", {}).get("reason", ""),
        "actual": actual,
        "verdict": verdict,
        "bug_id": case.get("known_defect", ""),
    }


def main():
    results = load_newman_results()
    specs = []
    for fname in sorted(os.listdir(TC_DIR)):
        if fname.endswith("_testcases.json"):
            with io.open(os.path.join(TC_DIR, fname), encoding="utf-8") as f:
                specs.append(json.load(f))
    if not specs:
        print("Khong tim thay IR nao trong %s" % TC_DIR, file=sys.stderr)
        return 1

    all_rows = OrderedDict()
    for spec in specs:
        rows = [flatten(spec, c, results) for c in spec["cases"] if not c.get("setup")]
        all_rows[spec["api_id"]] = (spec, rows)

        out_csv = os.path.join(TC_DIR, "%s_%s_testcases.csv"
                               % (spec["api_id"].lower().replace("-", ""), spec["fr"].lower()))
        with io.open(out_csv, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow([h for _, h in COLUMNS])
            for r in rows:
                w.writerow([r[k] for k, _ in COLUMNS])
        print("[CSV ] %-44s %3d test case" % (os.path.basename(out_csv), len(rows)))

    write_xlsx(all_rows)
    write_summary(all_rows, results)
    return 0


def write_xlsx(all_rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[XLSX] bo qua — chua cai openpyxl (pip install openpyxl)")
        return

    wb = Workbook()
    wb.remove(wb.active)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(color="FFFFFF", bold=True)
    fail_fill = PatternFill("solid", fgColor="FCE4E4")
    pass_fill = PatternFill("solid", fgColor="E7F4E7")
    human_fill = PatternFill("solid", fgColor="FFF4CE")

    for api_id, (spec, rows) in all_rows.items():
        ws = wb.create_sheet("%s %s" % (api_id, spec["fr"]))
        ws.append([h for _, h in COLUMNS])
        for cell in ws[1]:
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for r in rows:
            ws.append([r[k] for k, _ in COLUMNS])
        vcol = [k for k, _ in COLUMNS].index("verdict") + 1
        scol = [k for k, _ in COLUMNS].index("source") + 1
        for i in range(2, ws.max_row + 1):
            v = ws.cell(row=i, column=vcol).value
            if v == "FAIL":
                ws.cell(row=i, column=vcol).fill = fail_fill
            elif v == "PASS":
                ws.cell(row=i, column=vcol).fill = pass_fill
            if ws.cell(row=i, column=scol).value == "HUMAN":
                ws.cell(row=i, column=scol).fill = human_fill
        widths = {"Test Case ID": 15, "Tieu de test case": 52, "Ky thuat": 17,
                  "Tham so": 16, "Phan vung / Gia tri bien": 34,
                  "Du lieu gui len": 40, "Ket qua ky vong (theo dac ta)": 46,
                  "Ly do audit / hieu chinh": 60, "Ket qua chay that (Newman)": 44,
                  "Nhom": 40, "Endpoint": 26}
        for j, (_, h) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(j)].width = widths.get(h, 13)
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

    out = os.path.join(TC_DIR, "TESTCASES_23127195.xlsx")
    wb.save(out)
    print("[XLSX] %s" % os.path.basename(out))


def write_summary(all_rows, results):
    lines = ["# Test Summary — HW06 · 23127195", "",
             "> Sinh tu dong boi `scripts/export_testcases.py`. "
             "Cot *Ket qua chay* lay tu bao cao Newman moi nhat trong `newman/`.", ""]

    for key in ("api1", "api2", "api3"):
        f = results.get("__file__" + key)
        if f:
            lines.append("- Nguon ket qua **%s**: `newman/%s`" % (key.upper(), f[0]))
    lines += ["", "## Tong hop theo API", "",
              "| API | FR | Test case | AI sinh | SV tu them | PASS | FAIL | Bug lien quan |",
              "|---|---|---|---|---|---|---|---|"]

    tot = Counter()
    bugs_all = set()
    for api_id, (spec, rows) in all_rows.items():
        c = Counter(r["verdict"] for r in rows)
        s = Counter(r["source"] for r in rows)
        bugs = sorted({r["bug_id"] for r in rows if r["bug_id"]})
        bugs_all |= set(bugs)
        lines.append("| %s | %s | %d | %d | %d | %d | %d | %d |" % (
            api_id, spec["fr"], len(rows), s["AI"], s["HUMAN"],
            c["PASS"], c["FAIL"], len(bugs)))
        tot["cases"] += len(rows); tot["ai"] += s["AI"]; tot["human"] += s["HUMAN"]
        tot["pass"] += c["PASS"]; tot["fail"] += c["FAIL"]
    lines.append("| **Tong** | | **%d** | **%d** | **%d** | **%d** | **%d** | **%d** |" % (
        tot["cases"], tot["ai"], tot["human"], tot["pass"], tot["fail"], len(bugs_all)))

    lines += ["", "## Phan bo theo ky thuat kiem thu", "",
              "| API | Domain partition | State transition | Security | Schema |", "|---|---|---|---|---|"]
    for api_id, (spec, rows) in all_rows.items():
        t = Counter(r["technique"] for r in rows)
        lines.append("| %s | %d | %d | %d | %d |" % (
            api_id, t["domain-partition"], t["state-transition"], t["security"], t["schema"]))

    lines += ["", "## Ket qua audit (human review tren test case do AI sinh)", "",
              "| API | VALID | INCOMPLETE (da hieu chinh) | INVALID (da loai/sua) |", "|---|---|---|---|"]
    for api_id, (spec, rows) in all_rows.items():
        a = Counter(r["audit_label"] for r in rows if r["source"] == "AI")
        lines.append("| %s | %d | %d | %d |" % (api_id, a["VALID"], a["INCOMPLETE"], a["INVALID"]))

    out = os.path.join(TC_DIR, "TEST_SUMMARY.md")
    io.open(out, "w", encoding="utf-8", newline="").write("\n".join(lines) + "\n")
    print("[MD  ] %s" % os.path.basename(out))


if __name__ == "__main__":
    sys.exit(main())
