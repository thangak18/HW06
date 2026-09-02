#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""tc_to_excel.py - Gộp các CSV test case thành 1 file Excel có sheet Summary.

  python3 tc_to_excel.py --csv testcases/API-1_final.csv testcases/API-2_final.csv \
      testcases/API-3_final.csv --out testcases/23127060_HW06_testcases.xlsx

Cần openpyxl:  pip install openpyxl
Nếu không cài được openpyxl, dùng --fallback-html để xuất file .xls dạng HTML table
(Excel vẫn mở được), rồi Save As .xlsx bằng tay.
"""
import argparse
import csv
import os
from collections import Counter, defaultdict


def read_csv(path):
    with open(path, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def sheet_name(path):
    b = os.path.basename(path)
    for suf in ("_final.csv", "_audited.csv", "_generated.csv", ".csv"):
        if b.endswith(suf):
            return b[: -len(suf)][:31]
    return b[:31]


def build_summary(all_rows):
    """all_rows: dict {sheet: rows} -> list of summary rows"""
    out = []
    header = ["API", "Tong case", "AI sinh", "Tu bo sung", "VALID", "INVALID",
              "INCOMPLETE", "DOM", "STA", "SEC", "SCH", "@contract", "@bug",
              "P0", "SEC phu", "Ghi chu"]
    out.append(header)
    tot = Counter()
    for name, rows in all_rows.items():
        c_src = Counter(r.get("Source", "") for r in rows)
        c_lab = Counter((r.get("Audit_Label") or "").strip() for r in rows)
        c_cat = Counter(r.get("Category", "") for r in rows)
        c_tag = Counter(r.get("Tag", "") for r in rows)
        c_pri = Counter(r.get("Priority", "") for r in rows)
        secs = set(r.get("SEC_Ref") for r in rows)
        need = ["SEC-0%d" % i for i in range(1, 8)]
        missing = [s for s in need if s not in secs]
        note = ""
        if c_src.get("AI", 0) < 35:
            note += "THIEU case AI (<35). "
        if c_src.get("HUMAN", 0) < 5:
            note += "THIEU case tu bo sung (<5). "
        # Không phải mã SEC nào cũng áp dụng được cho mọi API: SEC-01 (lưu trữ mật khẩu) và
        # SEC-07 (vong doi OTP) chi lien quan den API-1. Yeu cau "du 7 ma cho MOI API" la bat
        # khả thi, và chính nó ép phải gán bừa mã SEC (xem report/03_audit.md mục 8). Chỉ tiêu
        # đúng là đủ 7 mã trên TOÀN BỘ suite; ở đây chỉ ghi nhận, không báo thiếu.
        if missing:
            note += "Khong ap dung: " + ",".join(missing) + ". "
        out.append([
            name, len(rows), c_src.get("AI", 0), c_src.get("HUMAN", 0),
            c_lab.get("VALID", 0), c_lab.get("INVALID", 0), c_lab.get("INCOMPLETE", 0),
            c_cat.get("DOM", 0), c_cat.get("STA", 0), c_cat.get("SEC", 0), c_cat.get("SCH", 0),
            c_tag.get("@contract", 0), c_tag.get("@bug", 0), c_pri.get("P0", 0),
            "%d/7 ap dung" % (7 - len(missing)), note.strip() or "OK",
        ])
        tot["total"] += len(rows)
        tot["ai"] += c_src.get("AI", 0)
        tot["hu"] += c_src.get("HUMAN", 0)
        tot["v"] += c_lab.get("VALID", 0)
        tot["i"] += c_lab.get("INVALID", 0)
        tot["n"] += c_lab.get("INCOMPLETE", 0)
        for k in ("DOM", "STA", "SEC", "SCH"):
            tot[k] += c_cat.get(k, 0)
        tot["c"] += c_tag.get("@contract", 0)
        tot["b"] += c_tag.get("@bug", 0)
        tot["p0"] += c_pri.get("P0", 0)
    out.append(["TONG", tot["total"], tot["ai"], tot["hu"], tot["v"], tot["i"], tot["n"],
                tot["DOM"], tot["STA"], tot["SEC"], tot["SCH"], tot["c"], tot["b"],
                tot["p0"], "", ""])
    return out


def write_xlsx(all_rows, summary, out):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF")

    for r in summary:
        ws.append(r)
    for c in ws[1]:
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for i in range(1, len(summary[0]) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions[get_column_letter(len(summary[0]))].width = 40

    cat_fill = {
        "DOM": PatternFill("solid", fgColor="E2EFDA"),
        "STA": PatternFill("solid", fgColor="DDEBF7"),
        "SEC": PatternFill("solid", fgColor="FCE4D6"),
        "SCH": PatternFill("solid", fgColor="FFF2CC"),
    }

    for name, rows in all_rows.items():
        s = wb.create_sheet(name)
        if not rows:
            continue
        cols = list(rows[0].keys())
        s.append(cols)
        for c in s[1]:
            c.fill = head_fill
            c.font = head_font
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        for r in rows:
            s.append([r.get(c, "") for c in cols])
        s.freeze_panes = "A2"
        s.auto_filter.ref = s.dimensions
        widths = {"TC_ID": 20, "Title": 55, "Endpoint": 38, "Request_Body": 45,
                  "Expected_Assertions": 45, "Preconditions": 30, "Audit_Note": 40,
                  "Why_AI_Missed": 40}
        for i, cname in enumerate(cols, 1):
            s.column_dimensions[get_column_letter(i)].width = widths.get(cname, 14)
        ci = cols.index("Category") + 1 if "Category" in cols else None
        if ci:
            for row_i in range(2, len(rows) + 2):
                v = s.cell(row=row_i, column=ci).value
                if v in cat_fill:
                    for cc in range(1, len(cols) + 1):
                        s.cell(row=row_i, column=cc).fill = cat_fill[v]

    d = os.path.dirname(os.path.abspath(out))
    if d:
        os.makedirs(d, exist_ok=True)
    wb.save(out)


def write_html(all_rows, summary, out):
    parts = ["<html><head><meta charset='utf-8'></head><body>"]
    parts.append("<h2>Summary</h2><table border=1>")
    for r in summary:
        parts.append("<tr>" + "".join("<td>%s</td>" % str(x) for x in r) + "</tr>")
    parts.append("</table>")
    for name, rows in all_rows.items():
        parts.append("<h2>%s</h2><table border=1>" % name)
        if rows:
            cols = list(rows[0].keys())
            parts.append("<tr>" + "".join("<th>%s</th>" % c for c in cols) + "</tr>")
            for r in rows:
                parts.append("<tr>" + "".join(
                    "<td>%s</td>" % str(r.get(c, "")).replace("<", "&lt;") for c in cols) + "</tr>")
        parts.append("</table>")
    parts.append("</body></html>")
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(parts))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", nargs="+", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--fallback-html", action="store_true")
    a = ap.parse_args()

    all_rows = {}
    for p in a.csv:
        all_rows[sheet_name(p)] = read_csv(p)
    summary = build_summary(all_rows)

    if a.fallback_html:
        out = os.path.splitext(a.out)[0] + ".xls"
        write_html(all_rows, summary, out)
        print("Da ghi (HTML table, Excel mo duoc): %s" % out)
        return

    try:
        write_xlsx(all_rows, summary, a.out)
    except ImportError:
        print("[CANH BAO] chua co openpyxl. Chay: pip install openpyxl")
        print("[CANH BAO] tam thoi xuat ban HTML thay the.")
        out = os.path.splitext(a.out)[0] + ".xls"
        write_html(all_rows, summary, out)
        print("Da ghi: %s" % out)
        return

    for r in summary[1:]:
        print("  %-22s tong=%-4s AI=%-4s HUMAN=%-3s  %s" % (r[0], r[1], r[2], r[3], r[-1]))
    print("Da ghi Excel: %s (%d sheet du lieu + 1 sheet Summary)" % (a.out, len(all_rows)))


if __name__ == "__main__":
    main()
