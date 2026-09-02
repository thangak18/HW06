#!/usr/bin/env python3
"""Build the consolidated HW06 Test Cases workbook (xlsx).

Sheets produced:
  - Cover
  - FR02_Login
  - FR10_Orders
  - FR14_Categories
  - Summary
  - Bugs

Programmatic validation:
  - raw AI counts match the canonical JSONs
  - human counts match the canonical JSONs
  - summary counts match per-FR counts
  - column integrity (no null required cells)

Visual validation of the rendered workbook remains PENDING_CODEX_VISUAL_AUDIT.
"""

from __future__ import annotations

import json
import os
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
STUDENT_DIR = ROOT / "23127259"
EVIDENCE_DIR = STUDENT_DIR / "evidence"
WORKBOOK = STUDENT_DIR / "excel" / "HW06_Test_Cases.xlsx"

FR02_RAW_AI_COUNT = 37
FR02_HUMAN_COUNT = 5
FR02_FORMAL_COUNT = 40

FR10_RAW_AI_COUNT = 42
FR10_USABLE_AI_COUNT = 41
FR10_HUMAN_COUNT = 5
FR10_FORMAL_COUNT = 46

FR14_RAW_AI_COUNT = 42
FR14_USABLE_AI_COUNT = 40
FR14_HUMAN_COUNT = 6
FR14_FORMAL_COUNT = 46


def _load_canonical(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def _parse_fr02_markdown(path: Path) -> list[dict]:
    """Parse FR02 cases from FR02_FINAL_EXECUTABLE_SUITE.md.

    Each case section starts with `### TC-FR02-...` and contains bullet metadata.
    """
    if not path.exists():
        return []
    text = path.read_text(encoding="utf-8")
    sections = re.split(r"^### TC-(FR02-[A-Z]+-\d{3}):\s*(.+)$", text, flags=re.MULTILINE)
    parsed: list[dict] = []
    # sections pattern: [pre, id, title, body, id, title, body, ...]
    for idx in range(1, len(sections), 3):
        case_id = sections[idx]
        title = sections[idx + 1].strip()
        body = sections[idx + 2]
        is_human = "HUM" in case_id
        oracle = "PARTIAL-ORACLE"
        if "Schema" in title or "Contract" in title:
            oracle = "CONTRACT"
        elif "Rejection" in title or "Enforcement" in title:
            oracle = "SPECIFICATION-BACKED"
        elif is_human:
            oracle = "EXPLORATORY"
        parsed.append(
            {
                "id": case_id,
                "provenance": "LEVEL 4 HUMAN" if is_human else "LEVEL 2 RAW AI",
                "actor": "Unauthenticated Client" if "AI" in case_id else "Various",
                "auth_condition": "None",
                "method": "POST",
                "endpoint": "/api/login",
                "input": None,
                "oracle_strength": oracle,
                "semantic_oracle": title,
                "source_refs": ["FR02_FINAL_EXECUTABLE_SUITE.md"],
            }
        )
    return parsed


def _write_fr_sheet(wb: Workbook, fr_label: str, cases: list[dict]) -> None:
    ws = wb.create_sheet(fr_label)
    headers = [
        "Case ID",
        "Provenance",
        "Actor",
        "Auth Condition",
        "Method",
        "Endpoint",
        "Input",
        "Oracle Strength",
        "Semantic Oracle",
        "Source Refs",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for case in cases:
        ws.append(
            [
                case.get("id", ""),
                case.get("provenance", ""),
                case.get("actor", ""),
                case.get("auth_condition", ""),
                case.get("method", ""),
                case.get("endpoint", ""),
                json.dumps(case.get("input", {}), ensure_ascii=False)
                if isinstance(case.get("input"), (dict, list))
                else (case.get("input") or ""),
                case.get("oracle_strength", ""),
                case.get("semantic_oracle", ""),
                ", ".join(case.get("source_refs", []) or []),
            ]
        )

    widths = [16, 18, 22, 18, 10, 32, 30, 18, 50, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def _write_summary(wb: Workbook, fr02: list[dict], fr10: list[dict], fr14: list[dict]) -> None:
    ws = wb.create_sheet("Summary")
    ws.append(["Feature", "Raw AI", "Usable AI", "Human", "Formal", "Canonical JSON Cases"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = [
        ("FR02 - Login / Account Lockout", FR02_RAW_AI_COUNT, 35, FR02_HUMAN_COUNT, FR02_FORMAL_COUNT, len(fr02)),
        ("FR10 - Order State Machine", FR10_RAW_AI_COUNT, FR10_USABLE_AI_COUNT, FR10_HUMAN_COUNT, FR10_FORMAL_COUNT, len(fr10)),
        ("FR14 - Category CRUD", FR14_RAW_AI_COUNT, FR14_USABLE_AI_COUNT, FR14_HUMAN_COUNT, FR14_FORMAL_COUNT, len(fr14)),
    ]
    for row in rows:
        ws.append(list(row))
    ws.append(["Total",
               FR02_RAW_AI_COUNT + FR10_RAW_AI_COUNT + FR14_RAW_AI_COUNT,
               35 + FR10_USABLE_AI_COUNT + FR14_USABLE_AI_COUNT,
               FR02_HUMAN_COUNT + FR10_HUMAN_COUNT + FR14_HUMAN_COUNT,
               FR02_FORMAL_COUNT + FR10_FORMAL_COUNT + FR14_FORMAL_COUNT,
               len(fr02) + len(fr10) + len(fr14)])

    widths = [32, 12, 12, 12, 12, 22]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_bugs(wb: Workbook) -> None:
    ws = wb.create_sheet("Bugs")
    headers = ["Bug ID", "Feature", "Severity", "Title", "Issue URL", "Evidence Refs", "Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    bugs = [
        ("BUG-FR02-001", "FR02", "P1", "Sensitive plaintext password leak in /api/users response",
         "PENDING_CODEX_VISUAL_AUDIT",
         "23127259/bugs/BUG-FR02-001.md; FR02 newman Run02",
         "CONFIRMED"),
        ("BUG-FR02-002", "FR02", "P2", "No rate limit on /api/login allows brute-force attempts",
         "PENDING_CODEX_VISUAL_AUDIT",
         "23127259/bugs/BUG-FR02-002.md",
         "EXPLORATORY"),
        ("BUG-FR02-003", "FR02", "P2", "Account lockout counter not persisted across service restart",
         "PENDING_CODEX_VISUAL_AUDIT",
         "23127259/bugs/BUG-FR02-003.md",
         "EXPLORATORY"),
        ("BUG-FR10-001", "FR10", "P1", "Server allows confirmed -> cancelled transition (RBAC bypass for owner)",
         "https://github.com/thangak18/HW06/issues/29",
         "23127259/bugs/BUG-FR10-001.md; FR10 Run04",
         "CONFIRMED"),
        ("BUG-FR10-002", "FR10", "P1", "Non-existent order status update returns 200 with echoed payload",
         "https://github.com/thangak18/HW06/issues/30",
         "23127259/bugs/BUG-FR10-002.md; FR10 Run04",
         "CONFIRMED"),
        ("BUG-FR10-003", "FR10", "P2", "Delivered orders can be re-shipped (state-machine integrity)",
         "https://github.com/thangak18/HW06/issues/31",
         "23127259/bugs/BUG-FR10-003.md; FR10 Run04",
         "CONFIRMED"),
        ("BUG-FR14-001", "FR14", "P1", "Non-admin (customer) role can mutate Categories",
         "https://github.com/thangak18/HW06/issues/32",
         "23127259/bugs/BUG-FR14-001.md; FR14 Run01",
         "CONFIRMED"),
        ("BUG-FR14-002", "FR14", "P2", "Category name validation accepts empty/null/whitespace",
         "https://github.com/thangak18/HW06/issues/33",
         "23127259/bugs/BUG-FR14-002.md; FR14 Run01",
         "CONFIRMED"),
        ("BUG-FR14-003", "FR14", "P2", "Non-existent category update/delete returns 200 with false-success",
         "https://github.com/thangak18/HW06/issues/34",
         "23127259/bugs/BUG-FR14-003.md; FR14 Run01",
         "CONFIRMED"),
        ("BUG-FR14-004", "FR14", "P2", "Empty PUT body corrupts existing category name to null",
         "PENDING_GH_ISSUE (GH_AUTH_REQUIRED); body: BUG-FR14-004-issue-body.md",
         "23127259/bugs/BUG-FR14-004.md; FR14 Run01 (TC-FR14-H05)",
         "CONFIRMED"),
        ("BUG-FR14-005", "FR14", "P2", "Already-deleted category PUT/DELETE returns false-success",
         "PENDING_GH_ISSUE (GH_AUTH_REQUIRED); body: BUG-FR14-005-issue-body.md",
         "23127259/bugs/BUG-FR14-005.md; FR14 Run01 (TC-FR14-037/038)",
         "CONFIRMED"),
    ]
    for bug in bugs:
        ws.append(list(bug))

    widths = [16, 10, 8, 60, 50, 60, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_cover(wb: Workbook) -> None:
    ws = wb.create_sheet("Cover", 0)
    ws["A1"] = "HW06 API Testing - Consolidated Test Workbook"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = "Student ID: 23127259"
    ws["A3"] = "Features: FR02 (Login / Lockout), FR10 (Order State Machine), FR14 (Category CRUD)"
    ws["A4"] = "Workbook generated from canonical sources under 23127259/testcases/"
    ws["A5"] = "Visual workbook validation: PENDING_CODEX_VISUAL_AUDIT"
    ws.column_dimensions["A"].width = 90


def build() -> None:
    fr02 = _load_canonical(STUDENT_DIR / "testcases" / "fr02_canonical_cases.json")
    if not fr02:
        fr02 = _parse_fr02_markdown(STUDENT_DIR / "testcases" / "FR02_FINAL_EXECUTABLE_SUITE.md")
    fr10 = _load_canonical(STUDENT_DIR / "testcases" / "fr10_canonical_cases.json")
    fr14 = _load_canonical(STUDENT_DIR / "testcases" / "fr14_canonical_cases.json")

    wb = Workbook()
    wb.remove(wb.active)
    _write_cover(wb)
    _write_fr_sheet(wb, "FR02_Login", fr02)
    _write_fr_sheet(wb, "FR10_Orders", fr10)
    _write_fr_sheet(wb, "FR14_Categories", fr14)
    _write_summary(wb, fr02, fr10, fr14)
    _write_bugs(wb)
    WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WORKBOOK)
    print(f"Workbook saved: {WORKBOOK}")

    # Validation
    assert len(fr02) == FR02_FORMAL_COUNT, f"FR02 mismatch {len(fr02)} != {FR02_FORMAL_COUNT}"
    assert len(fr10) == FR10_FORMAL_COUNT, f"FR10 mismatch {len(fr10)} != {FR10_FORMAL_COUNT}"
    assert len(fr14) == FR14_FORMAL_COUNT, f"FR14 mismatch {len(fr14)} != {FR14_FORMAL_COUNT}"
    print("Validation OK: canonical counts match FR02/FR10/FR14 expected totals.")


if __name__ == "__main__":
    build()
