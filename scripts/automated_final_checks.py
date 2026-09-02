#!/usr/bin/env python3
"""HW06 automated non-visual final checks.

Performs programmatic verification of:
  - required file existence
  - canonical counts (raw AI / human / formal)
  - X-Student-Id static coverage
  - hostnames in textual evidence
  - duplicate IDs
  - workbook integrity
  - CI workflow YAML structure
  - secret hygiene in text artifacts
  - bug evidence / GitHub issue references

This script intentionally avoids image-content analysis. It only
inspects metadata (filename, size, SHA-256) for visual files.

Exit code 0 on success; non-zero on any failure.
"""

from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
STUDENT_DIR = ROOT / "23127259"
EV_FR02 = STUDENT_DIR / "newman" / "fr02"
EV_FR10 = STUDENT_DIR / "evidence" / "fr10" / "newman"
EV_FR14 = STUDENT_DIR / "evidence" / "fr14" / "newman"
COLLECTION_DIR = STUDENT_DIR / "postman" / "collections"
ENV_DIR = STUDENT_DIR / "postman" / "environments"

REQUIRED_FILES = [
    "23127259/README.md",
    "23127259/docs/00_MAIN_REPORT.md",
    "23127259/docs/test_generator.md",
    "23127259/docs/AI_TEST_GENERATOR_DIAGRAM_SPEC.md",
    "23127259/ai/AI_AUDIT_REPORT.md",
    "23127259/ai/AI_CRITIQUE.md",
    "23127259/excel/HW06_Test_Cases.xlsx",
    "23127259/audit/HW06_REQUIREMENTS_COMPLIANCE_MATRIX.md",
    "23127259/audit/FINAL_3_FEATURE_COMPLIANCE_AUDIT.md",
    "23127259/audit/CODEX_VISUAL_HANDOFF.md",
    "23127259/evidence/FINAL_EVIDENCE_MANIFEST.md",
    "23127259/ci/CI_CD_REPORT.md",
    "scripts/run_fr10_newman.sh",
    "scripts/run_fr14_newman.sh",
    "scripts/build_excel_workbook.py",
    "scripts/build_fr14_collection.py",
    "scripts/validate_fr14_collection.py",
    "scripts/sanitize_fr14_artifacts.py",
    ".github/workflows/hw06-23127259-api-tests.yml",
    ".github/workflows/hw06-deliberate-red.yml",
]

STUDENT_ID = "23127259"
JWT_PATTERN = re.compile(r"eyJ[A-Za-z0-9_-]{10,}\.[A-Za-z0-9_-]{10,}\.[A-Za-z0-9_-]{10,}")
RESOLVED_BEARER = re.compile(r"Bearer\s+eyJ[A-Za-z0-9_-]{10,}\.[A-Za-z0-9_-]{10,}\.[A-Za-z0-9_-]{10,}")


def _sha256(path: Path) -> str:
    if not path.exists() or path.stat().st_size == 0:
        return ""
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _print_section(title: str) -> None:
    print("\n" + "=" * 60)
    print(title)
    print("=" * 60)


def check_required_files() -> list[str]:
    errors = []
    _print_section("Required Files")
    for rel in REQUIRED_FILES:
        path = ROOT / rel
        ok = path.exists() and path.stat().st_size > 0
        print(f"  [{ 'OK' if ok else 'MISSING' }] {rel}")
        if not ok:
            errors.append(f"missing or empty: {rel}")
    return errors


def check_canonical_counts() -> list[str]:
    errors = []
    _print_section("Canonical Counts")
    fr10 = json.loads((STUDENT_DIR / "testcases" / "fr10_canonical_cases.json").read_text())
    fr14 = json.loads((STUDENT_DIR / "testcases" / "fr14_canonical_cases.json").read_text())
    fr10_human = [c for c in fr10 if "HUM" in c.get("id", "")]
    fr14_human = [c for c in fr14 if "HUM" in c.get("id", "")]
    print(f"  FR10 canonical count: {len(fr10)} (expected 46)")
    print(f"  FR10 human count: {len(fr10_human)} (expected >= 5)")
    print(f"  FR14 canonical count: {len(fr14)} (expected 46)")
    print(f"  FR14 human count: {len(fr14_human)} (expected >= 5)")
    if len(fr10) != 46:
        errors.append(f"FR10 count mismatch: {len(fr10)} != 46")
    if len(fr14) != 46:
        errors.append(f"FR14 count mismatch: {len(fr14)} != 46")
    if len(fr10_human) < 5:
        errors.append(f"FR10 human count {len(fr10_human)} < 5")
    if len(fr14_human) < 5:
        errors.append(f"FR14 human count {len(fr14_human)} < 5")
    fr10_ids = Counter(c["id"] for c in fr10)
    fr14_ids = Counter(c["id"] for c in fr14)
    dup10 = {k: v for k, v in fr10_ids.items() if v > 1}
    dup14 = {k: v for k, v in fr14_ids.items() if v > 1}
    if dup10:
        errors.append(f"FR10 duplicate IDs: {dup10}")
    if dup14:
        errors.append(f"FR14 duplicate IDs: {dup14}")
    if not dup10 and not dup14:
        print("  No duplicate formal IDs.")
    return errors


def check_x_student_id_static() -> list[str]:
    errors = []
    _print_section("X-Student-Id Static Coverage")
    for collection_path in COLLECTION_DIR.glob("*.json"):
        text = collection_path.read_text()
        student_id_count = text.count(STUDENT_ID)
        print(f"  {collection_path.name}: {student_id_count} occurrences")
        if student_id_count == 0:
            errors.append(f"{collection_path.name} missing X-Student-Id")
    for env_path in ENV_DIR.glob("*.json"):
        text = env_path.read_text()
        student_id_count = text.count(STUDENT_ID)
        print(f"  {env_path.name}: {student_id_count} occurrences (env)")
    return errors


def check_workbook_integrity() -> list[str]:
    errors = []
    _print_section("Excel Workbook Integrity")
    try:
        from openpyxl import load_workbook
    except ImportError:
        errors.append("openpyxl not available")
        return errors
    workbook_path = STUDENT_DIR / "excel" / "HW06_Test_Cases.xlsx"
    if not workbook_path.exists():
        errors.append("Excel workbook missing")
        return errors
    wb = load_workbook(workbook_path)
    expected_sheets = {"Cover", "FR02_Login", "FR10_Orders", "FR14_Categories", "Summary", "Bugs"}
    actual_sheets = set(wb.sheetnames)
    missing = expected_sheets - actual_sheets
    if missing:
        errors.append(f"workbook missing sheets: {missing}")
    print(f"  Sheets: {wb.sheetnames}")
    if "FR10_Orders" in wb.sheetnames:
        rows = wb["FR10_Orders"].max_row
        print(f"  FR10_Orders rows: {rows} (expected 47 = 1 header + 46 cases)")
        if rows != 47:
            errors.append(f"FR10_Orders row count wrong: {rows}")
    if "FR14_Categories" in wb.sheetnames:
        rows = wb["FR14_Categories"].max_row
        print(f"  FR14_Categories rows: {rows} (expected 47)")
        if rows != 47:
            errors.append(f"FR14_Categories row count wrong: {rows}")
    return errors


def check_ci_workflow_yaml() -> list[str]:
    errors = []
    _print_section("CI Workflow YAML")
    for wf in ["hw06-23127259-api-tests.yml", "hw06-deliberate-red.yml"]:
        path = ROOT / ".github" / "workflows" / wf
        if not path.exists():
            errors.append(f"missing workflow: {wf}")
            continue
        text = path.read_text()
        required_tokens = ["name:", "on:", "jobs:", "runs-on:", "steps:", "STUDENT_ID"]
        missing = [t for t in required_tokens if t not in text]
        if missing:
            errors.append(f"{wf} missing tokens: {missing}")
        else:
            print(f"  {wf}: OK")
    return errors


def check_secret_hygiene() -> list[str]:
    errors = []
    _print_section("Secret Hygiene (text artifacts)")
    text_roots = [
        STUDENT_DIR / "evidence",
        STUDENT_DIR / "postman",
        STUDENT_DIR / "audit",
        STUDENT_DIR / "docs",
        STUDENT_DIR / "bugs",
        STUDENT_DIR / "ci",
        ROOT / "scripts",
    ]
    issues = []
    public_safe_root = STUDENT_DIR / "evidence"
    historical_evidence_roots = [
        STUDENT_DIR / "evidence" / "fr10" / "confirmation",
        STUDENT_DIR / "evidence" / "fr10" / "newman",
    ]
    historical_evidence_names = (
        "FR10-confirmation",
        "FR10-bug-evidence",
        "FR10-run02",
        "FR10-run03",
    )
    for root in text_roots:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if not path.is_file():
                continue
            if path.suffix in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".pdf", ".xlsx"}:
                continue
            if path.stat().st_size > 5_000_000:
                continue
            try:
                text = path.read_text(errors="ignore")
            except Exception:
                continue
            is_public_safe = "public-safe" in path.parts
            is_immutable_historical = (
                any(path.is_relative_to(hist) for hist in historical_evidence_roots)
                and path.name.startswith(historical_evidence_names)
            )
            if JWT_PATTERN.search(text):
                if "tampered" in text.lower() and "FR14" in str(path):
                    continue
                if "build_fr14_collection" in str(path):
                    continue
                if "validate_fr10_tampered" in str(path):
                    continue
                if is_public_safe and RESOLVED_BEARER.search(text):
                    issues.append(f"resolved Bearer in public-safe: {path.relative_to(ROOT)}")
                    continue
                if is_immutable_historical:
                    continue
                issues.append(f"JWT pattern in {path.relative_to(ROOT)}")
    for issue in issues:
        print(f"  [SECRET] {issue}")
        errors.append(issue)
    if not issues:
        print("  No raw JWT/Bearer secrets detected in non-public-safe text artifacts.")
    return errors


def check_newman_artifact_integrity() -> list[str]:
    errors = []
    _print_section("Newman Artifact Integrity")
    expected = [
        (EV_FR02 / "FR02-run-03.json", "FR02-run-03"),
        (EV_FR02 / "FR02-run-03.html", "FR02-run-03"),
        (EV_FR10 / "FR10-run04-exitcode.txt", "FR10-run04"),
        (EV_FR14 / "FR14-run01-exitcode.txt", "FR14-run01 (current canonical until Run05 replaces)"),
    ]
    for path, label in expected:
        if not path.exists():
            errors.append(f"missing: {path.relative_to(ROOT)} ({label})")
            continue
        size = path.stat().st_size
        print(f"  {path.relative_to(ROOT)}: {size} bytes ({label})")
    return errors


def check_bug_evidence() -> list[str]:
    errors = []
    _print_section("Bug Evidence")
    bug_root = STUDENT_DIR / "bugs"
    bug_files = list(bug_root.glob("BUG-*.md")) + list((bug_root / "issues").glob("BUG-*.md"))
    expected_ids = {"FR02-001", "FR02-002", "FR02-003",
                    "FR10-001", "FR10-002", "FR10-003",
                    "FR14-001", "FR14-002", "FR14-003", "FR14-004", "FR14-005"}
    found_ids = {f.stem.replace("BUG-", "") for f in bug_files}
    found_ids = {fid for fid in found_ids if "-issue-body" not in fid}
    missing = expected_ids - found_ids
    if missing:
        errors.append(f"missing bug reports: {missing}")
    print(f"  Found: {sorted(found_ids)}")
    return errors


def main() -> int:
    errors: list[str] = []
    errors += check_required_files()
    errors += check_canonical_counts()
    errors += check_x_student_id_static()
    errors += check_workbook_integrity()
    errors += check_ci_workflow_yaml()
    errors += check_secret_hygiene()
    errors += check_newman_artifact_integrity()
    errors += check_bug_evidence()

    _print_section("Result")
    if errors:
        print(f"FAILED with {len(errors)} error(s):")
        for e in errors:
            print(f"  - {e}")
        return 1
    print("ALL CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())